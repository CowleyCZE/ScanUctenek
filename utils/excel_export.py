"""
Excel export utility for receipt data
Exports receipts to Excel file with metadata as comments
"""

import pandas as pd
from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from typing import List, Dict, Any, Optional
import os
import logging
from .cell_mapping import get_cell_range, find_next_empty_cell

# Configure logging
logger = logging.getLogger(__name__)

def validate_receipt(receipt: Dict[str, Any]) -> bool:
    """
    Validuje data účtenky před exportem.
    
    Args:
        receipt: Slovník s daty účtenky
        
    Returns:
        bool: True pokud jsou data validní, False jinak
    """
    required_fields = ['merchant', 'date', 'total', 'currency', 'payment_method']
    
    try:
        # Kontrola povinných polí
        for field in required_fields:
            if field not in receipt:
                logger.error(f"Chybí povinné pole: {field}")
                return False
                
        # Validace datových typů
        if not isinstance(receipt['merchant'], str):
            logger.error("Obchodník musí být text")
            return False
            
        if not isinstance(receipt['date'], datetime):
            logger.error("Datum musí být datetime objekt")
            return False
            
        if not isinstance(receipt['total'], (int, float)):
            logger.error("Celková částka musí být číslo")
            return False
            
        if receipt['total'] < 0:
            logger.error("Celková částka nemůže být záporná")
            return False
            
        if receipt['currency'] not in ['EUR', 'CZK']:
            logger.error("Neplatná měna")
            return False
            
        if receipt['payment_method'] not in ['Kartou', 'Hotovost']:
            logger.error("Neplatný způsob platby")
            return False
            
        return True
        
    except Exception as e:
        logger.error(f"Chyba při validaci účtenky: {str(e)}")
        return False

def safe_path_join(*paths: str) -> str:
    """
    Bezpečné spojení cest k souborům.
    
    Args:
        *paths: Cesty k souborům
        
    Returns:
        Absolutní cesta k souboru
    """
    return os.path.abspath(os.path.join(*paths))

def export_to_excel(receipts: List[Dict[str, Any]], column_mapping: Dict[str, str], template_path: Optional[str] = None) -> BytesIO:
    """
    Exportuje účtenky do Excel souboru.
    
    Args:
        receipts: Seznam účtenek k exportu
        column_mapping: Mapování sloupců
        template_path: Volitelná cesta k šabloně
        
    Returns:
        BytesIO objekt s Excel souborem
    """
    import streamlit as st
    
    # Inicializace výstupního bufferu
    output = BytesIO()
    
    try:
        # Validace vstupních dat
        if not receipts:
            logger.warning("Žádné účtenky k exportu")
            return output
            
        # Validace každé účtenky
        valid_receipts = [r for r in receipts if validate_receipt(r)]
        if not valid_receipts:
            logger.error("Žádné validní účtenky k exportu")
            return output
            
        # Vytvoření nebo načtení workbooku
        if template_path and os.path.exists(template_path):
            try:
                workbook = openpyxl.load_workbook(template_path)
            except Exception as e:
                logger.error(f"Chyba při načítání šablony: {str(e)}")
                workbook = openpyxl.Workbook()
        else:
            workbook = openpyxl.Workbook()
            
        # Použití prvního listu nebo vytvoření nového
        if 'List1' in workbook.sheetnames:
            sheet = workbook['List1']
        else:
            sheet = workbook.active
            sheet.title = 'List1'
            
        # Nastavení hlaviček pro nový list
        if sheet['B11'].value is None:
            sheet['B11'] = 'EUR Karta'
            sheet['C11'] = 'EUR Hotově'
            sheet['D11'] = 'CZK Karta'
            sheet['E11'] = 'CZK Hotově'
            
            sheet['A12'] = 'Pohonné hmoty'
            sheet['A20'] = 'Mýtné'
            sheet['A27'] = 'Ubytování'
            sheet['A32'] = 'Ostatní'
            
        # Zpracování každé účtenky
        for receipt in valid_receipts:
            try:
                # Získání rozsahu buněk
                cell_range = get_cell_range(
                    receipt.get('purpose', 'Ostatní'),
                    receipt.get('currency', 'CZK'),
                    receipt.get('payment_method', 'Hotovost')
                )
                
                if not cell_range:
                    logger.warning(f"Nenalezen rozsah buněk pro účtenku: {receipt.get('merchant')}")
                    continue
                    
                # Hledání další prázdné buňky
                cell = find_next_empty_cell(sheet, cell_range)
                if not cell:
                    logger.warning(f"Žádná prázdná buňka v rozsahu pro účtenku: {receipt.get('merchant')}")
                    continue
                    
                # Zápis částky
                sheet[cell] = float(receipt['total'])
                
                # Přidání komentáře s detaily
                comment_text = (
                    f"Datum: {receipt['date'].strftime('%d.%m.%Y')}\n"
                    f"Obchodník: {receipt.get('merchant', '')}\n"
                    f"Číslo účtenky: {receipt.get('receipt_number', '')}"
                )
                comment = Comment(comment_text, "Detaily účtenky")
                sheet[cell].comment = comment
                
            except Exception as e:
                logger.error(f"Chyba při zpracování účtenky: {str(e)}")
                continue
                
        # Uložení workbooku
        workbook.save(output)
        output.seek(0)
        return output
        
    except Exception as e:
        logger.error(f"Chyba při vytváření Excel souboru: {str(e)}")
        raise
