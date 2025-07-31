import pytest
from datetime import datetime
import openpyxl

from utils.excel_export import export_to_excel, validate_receipt
from utils.cell_mapping import get_cell_range, find_next_empty_cell

# --- Testy pro validaci účtenky ---

@pytest.fixture
def valid_receipt_data():
    """Poskytuje validní data účtenky pro testy."""
    return {
        'merchant': 'Testovací Obchod',
        'date': datetime(2024, 7, 30),
        'total': 150.75,
        'currency': 'CZK',
        'payment_method': 'Kartou',
        'purpose': 'Pohonné hmoty',
        'receipt_number': '12345'
    }

def test_validate_receipt_valid(valid_receipt_data):
    """Testuje validaci s kompletními a správnými daty."""
    assert validate_receipt(valid_receipt_data) is True

def test_validate_receipt_invalid_missing_field(valid_receipt_data):
    """Testuje validaci, když chybí povinné pole."""
    del valid_receipt_data['total']
    assert validate_receipt(valid_receipt_data) is False

def test_validate_receipt_invalid_data_type(valid_receipt_data):
    """Testuje validaci s nesprávným datovým typem."""
    valid_receipt_data['total'] = 'sto padesát'
    assert validate_receipt(valid_receipt_data) is False

def test_validate_receipt_invalid_value(valid_receipt_data):
    """Testuje validaci s neplatnou hodnotou (např. záporná částka)."""
    valid_receipt_data['total'] = -100
    assert validate_receipt(valid_receipt_data) is False

    valid_receipt_data['total'] = 100
    valid_receipt_data['currency'] = 'USD' # Nepodporovaná měna
    assert validate_receipt(valid_receipt_data) is False

# --- Testy pro export do Excelu ---

def test_export_to_excel_single_receipt(valid_receipt_data):
    """Testuje export jedné účtenky a ověřuje obsah buňky a komentáře."""
    receipts = [valid_receipt_data]
    column_mapping = {
        'date': 'Datum', 'merchant': 'Obchodník', 'total': 'Celková částka',
        'payment_method': 'Způsob platby', 'purpose': 'Účel', 'currency': 'Měna'
    }

    excel_bytes = export_to_excel(receipts, column_mapping)
    assert excel_bytes is not None

    # Načtení workbooku z paměti pro ověření obsahu
    workbook = openpyxl.load_workbook(excel_bytes)
    sheet = workbook.active

    # Očekávaná buňka pro ('Pohonné hmoty', 'CZK', 'Kartou') je D12
    target_cell = 'D12'
    cell = sheet[target_cell]

    assert cell.value == valid_receipt_data['total']
    assert cell.comment is not None
    assert valid_receipt_data['merchant'] in cell.comment.text
    assert valid_receipt_data['receipt_number'] in cell.comment.text

def test_export_to_excel_multiple_receipts(valid_receipt_data):
    """Testuje export více účtenek a ověřuje, že se zapisují do správných buněk."""
    receipt1 = valid_receipt_data
    receipt2 = {
        'merchant': 'Jiný Obchod', 'date': datetime(2024, 7, 31), 'total': 2500.0,
        'currency': 'EUR', 'payment_method': 'Hotovost', 'purpose': 'Mýtné', 'receipt_number': '67890'
    }
    receipts = [receipt1, receipt2]
    column_mapping = {
        'date': 'Datum', 'merchant': 'Obchodník', 'total': 'Celková částka',
        'payment_method': 'Způsob platby', 'purpose': 'Účel', 'currency': 'Měna'
    }

    excel_bytes = export_to_excel(receipts, column_mapping)
    workbook = openpyxl.load_workbook(excel_bytes)
    sheet = workbook.active

    # Ověření první účtenky
    cell1 = sheet['D12'] # Pohonné hmoty, CZK, Kartou
    assert cell1.value == receipt1['total']

    # Ověření druhé účtenky
    cell2 = sheet['C20'] # Mýtné, EUR, Hotovost
    assert cell2.value == receipt2['total']
    assert receipt2['merchant'] in cell2.comment.text

def test_export_to_excel_bug_ubytovani_vs_bydleni(valid_receipt_data):
    """
    Tento test odhaluje chybu, kdy 'Ubytování' není správně mapováno,
    protože v CELL_MAPPINGS je 'Bydlení'.
    """
    receipt = valid_receipt_data
    receipt['purpose'] = 'Ubytování' # standardize_purpose vrací 'Ubytování'
    receipt['currency'] = 'EUR'
    receipt['payment_method'] = 'Kartou'

    # Očekáváme, že get_cell_range selže, protože 'Ubytování' není v mapování
    cell_range = get_cell_range(receipt['purpose'], receipt['currency'], receipt['payment_method'])
    assert cell_range is None

    # V důsledku toho by export neměl pro tuto účtenku nic zapsat
    excel_bytes = export_to_excel([receipt], {})
    workbook = openpyxl.load_workbook(excel_bytes)
    sheet = workbook.active

    # Očekávaná buňka B27 by měla být prázdná
    assert sheet['B27'].value is None
